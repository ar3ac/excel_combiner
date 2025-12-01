# leggere tutti i file excel nella cartella input

# combinarli in un unico file

# aggiungere colonna con nome del file

# salvare il risultato in output
from pathlib import Path
from openpyxl import load_workbook, Workbook
import csv
import configparser

config = configparser.ConfigParser()
config.read("config.ini")
input_folder = config["paths"]["input_folder"]
output_folder = config["paths"]["output_folder"]

Path(output_folder).mkdir(parents=True, exist_ok=True)


def get_excel_files(input_folder):
    input_path = Path(input_folder)
    excel_files = list(input_path.glob("*.xlsx"))
    filelist = [f.name for f in excel_files]
    print(f"Trovati {len(filelist)} file Excel:")
    for file in filelist:
        print(f"- {file}")
    return excel_files


def load_workbook_safe(file_path):

    try:
        workbook = load_workbook(filename=file_path, data_only=True)
        return workbook
    except Exception as e:
        print(f"Errore nel caricamento del file {file_path}: {e}")
        return None


def read_sheet_as_rows(workbook):
    sheet = workbook[workbook.sheetnames[0]]
    row_tuple = list(sheet.iter_rows(values_only=True))
    num_columns = len(row_tuple[0])
    num_rows = len(row_tuple)
    rows_list = [list(row) for row in row_tuple]
    return rows_list, num_columns, num_rows


def combine_rows(header, all_rows):
    combined = [header]
    for key, rows in all_rows.items():
        if key == "header":
            continue
        combined.extend(rows)
    return combined


def write_combined_excel(combined_rows, output_path):
    wb_final = Workbook()
    ws_final = wb_final.active
    for riga in combined_rows:
        ws_final.append(riga)
    wb_final.save(output_path)


def write_combined_csv(combined_rows, output_path):
    with open(output_path, mode="w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file, delimiter=";")
        writer.writerows(combined_rows)


if __name__ == "__main__":
    pathlist = get_excel_files(input_folder)
    rows_per_file = {}
    for file_path in pathlist:
        workbook = load_workbook_safe(file_path)
        if workbook is not None:
            print(f"Carico {file_path.name} ... OK, ", end="")
            rows, num_columns, num_rows = read_sheet_as_rows(workbook)
            print(f"lette {num_rows} righe e {num_columns} colonne.")
            if "header" not in rows_per_file:
                rows_per_file["header"] = rows[0] + ["source_file"]
            if rows[0] != rows_per_file["header"][:-1]:
                print(
                    f"Attenzione: l'header del file {file_path.name} è diverso dagli altri."
                )
            rows_with_filename = [row + [file_path.name] for row in rows[1:]]
            rows_per_file[file_path.name] = rows_with_filename
    header = rows_per_file.get("header", [])
    print(f"Header utilizzato: {header}")
    combined_rows = combine_rows(header, rows_per_file)
    print("File processati :", len(rows_per_file) - 1)
    print("Righe combinate (header escluso):", len(combined_rows) - 1)
    print("Righe combinate (header incluso):", len(combined_rows))
    write_combined_excel(combined_rows, output_folder + "combined.xlsx")

    print("Combinazione completata.")
    print(f"File generato: {output_folder}combined.xlsx")
    print(f"Totale righe: {len(combined_rows)} (header incluso)")

    write_combined_csv(combined_rows, output_folder + "combined.csv")
    print(f"File CSV generato: {output_folder}combined.csv")
    print(f"Totale righe: {len(combined_rows)} (header incluso)")
